from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
import os
from io import BytesIO
from datetime import datetime, timedelta

DATABASE_PATH = 'databases/Run Sheet Database.db'


region_colors = {
    'North Shore': 'F79646',
    'Quebec': '4F81BD',
    'Montreal': 'C0504D',
    'Ontario': '000000',
    'Sherbrooke': '00B0F0',
    'South Shore': '8064A2',
    'Beauce': 'FFC0CB',        # Pink instead of yellow
    'Drummond': '9BBB59',
}

region_layout = {
    'North Shore': ('A1', 'B1'),
    'Quebec': ('A20', 'B20'),
    'Montreal': ('A39', 'B39'),
    'Ontario': ('A58', 'B58'),
    'Drummond': ('K1', 'L1'),
    'Beauce': ('K20', 'L20'),
    'South Shore': ('K39', 'L39'),
    'Sherbrooke': ('K58', 'L58'),
}

headers = ["Customer ID", "Customer Name", "City", "Weight", "Skids", "Bundles", "Coils", "Closing Time", "Pickup"]
header_font = Font(bold=True, color="FFFFFF")
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def get_latest_runsheet_filename():
    today = datetime.today()
    if today.weekday() == 4:  # Friday
        run_date = today.replace(day=today.day + 3)
    else:
        run_date = today.replace(day=today.day + 1)

    filename = run_date.strftime("%A").upper() + f" RUN SHEET {run_date:%m.%d.%Y}.xlsx"
    folder = run_date.strftime("%B %Y")
    return filename, folder

def save_run_sheet_to_excel():
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM 'Run Sheet'")
    entries = cursor.fetchall()

    try:
        cursor.execute("SELECT * FROM Drivers")
        drivers = {row["Region"]: row["Name"] for row in cursor.fetchall()}
    except:
        drivers = {}

    connection.close()

    region_data = {region: [] for region in region_layout}
    for row in entries:
        region = row['Region']
        if region in region_data:
            region_data[region].append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Run Sheet"

    for region, (start_cell, driver_cell) in region_layout.items():
        start_col = ord(start_cell[0].upper()) - ord('A') + 1
        start_row = int(start_cell[1:])
        driver_col = ord(driver_cell[0].upper()) - ord('A') + 1

        region_cell = ws.cell(row=start_row, column=start_col, value=f"Region: {region}")
        region_cell.font = Font(bold=True, color=region_colors[region])
        region_cell.alignment = Alignment(horizontal="left")

        ws.cell(row=start_row, column=driver_col, value=drivers.get(region, "____________________"))

        for i, header in enumerate(headers):
            cell = ws.cell(row=start_row + 1, column=start_col + i, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type='solid')
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        region_rows = region_data[region]
        for r, row in enumerate(region_rows):
            for c, key in enumerate(headers):
                val = row[key]
                cell = ws.cell(row=start_row + 2 + r, column=start_col + c, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        for r in range(len(region_rows), 16):
            for c in range(len(headers)):
                cell = ws.cell(row=start_row + 2 + r, column=start_col + c)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        total_row = start_row + 18
        ws.cell(row=total_row, column=start_col + 2, value="Totals:").alignment = Alignment(horizontal="right")
        for offset, key in enumerate(["Weight", "Skids", "Bundles", "Coils"]):
            col = start_col + 3 + offset
            total = sum(float(r[key]) if r[key] not in [None, ""] else 0 for r in region_rows)
            cell = ws.cell(row=total_row, column=col, value=round(total))
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for col in range(1, 20):
            max_len = 0
            for row in range(1, 77):
                cell = ws.cell(row=row, column=col)
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = max(max_len + 2, 8)

    # ✅ Save in memory instead of disk
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    save_totals_to_db(region_data)
    return output

def save_totals_to_db(region_data):
    total_weight = total_skids = total_bundles = total_coils = 0

    def parse_num(val):
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0

    for customers in region_data.values():
        for entry in customers:
            total_weight += parse_num(entry['Weight'])
            total_skids += parse_num(entry['Skids'])
            total_bundles += parse_num(entry['Bundles'])
            total_coils += parse_num(entry['Coils'])

    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO RunSheetTotals (date, total_weight, total_skids, total_bundles, total_coils)
        VALUES (?, ?, ?, ?, ?)
    """, (datetime.today().strftime('%Y-%m-%d'), total_weight, total_skids, total_bundles, total_coils))
    conn.commit()
    conn.close()