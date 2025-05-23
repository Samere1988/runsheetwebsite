from flask import Blueprint, render_template, request, redirect, url_for, flash, send_from_directory,jsonify, send_file
import sqlite3
import os
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from .save_runsheet import save_run_sheet_to_excel, get_latest_runsheet_filename
from datetime import datetime, time, timedelta


runsheet_bp = Blueprint('runsheet', __name__)

DATABASE_PATH = 'databases/Run Sheet Database.db'

# -------------------
# View Run Sheet
# -------------------
# View Run Sheet
# -------------------
@runsheet_bp.route('/view_run_sheet')
def view_run_sheet():
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    # Fetch data
    cursor.execute("SELECT * FROM 'Run Sheet'")
    runsheet = cursor.fetchall()

    cursor.execute("SELECT * FROM Drivers")
    driver_rows = cursor.fetchall()
    connection.close()

    # Define regions
    regions = ['North Shore', 'Quebec', 'Montreal', 'Ontario', 'Drummond', 'Beauce', 'South Shore', 'Sherbrooke']
    region_data = {region: [] for region in regions}
    region_totals = {region: {"Weight": 0, "Skids": 0, "Bundles": 0, "Coils": 0} for region in regions}
    driver_info = {region: {"driver": "____________________"} for region in regions}

    # Helpers
    def safe_string(value):
        if value is None:
            return ''
        if isinstance(value, time):
            return value.strftime('%H:%M')
        return str(value).strip()

    def safe_int(value):
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0

    # Organize customer data
    for entry in runsheet:
        region = entry['Region']
        if region in region_data:
            region_data[region].append(entry)
            region_totals[region]["Weight"] += safe_int(entry['Weight'])
            region_totals[region]["Skids"] += safe_int(entry['Skids'])
            region_totals[region]["Bundles"] += safe_int(entry['Bundles'])
            region_totals[region]["Coils"] += safe_int(entry['Coils'])

    # Fill driver info from DB
    for row in driver_rows:
        region = row['Region']
        if region in driver_info:
            driver_info[region]['driver'] = row['Name']

    return render_template(
        'view_run_sheet.html',
        region_data=region_data,
        region_totals=region_totals,
        driver_info=driver_info
    )

# -------------------
# Add Customer
# -------------------
@runsheet_bp.route('/add_customer', methods=['GET', 'POST'])
def add_customer():
    customer_data = None
    customer_id = request.args.get('customer_id')

    if request.method == 'POST':
        form = request.form
        customer_id = form['customer_id']
        customer_name = form['customer_name']
        address = form['address']
        city = form['city']
        province = form['province']
        postal_code = form['postal_code']
        region = form['region']
        weight = int(form['weight'])
        skids = int(form['skids'])
        bundles = int(form['bundles'])
        coils = int(form['coils'])
        closing_time = form['closing_time']
        pickup = form['pickup']

        connection = sqlite3.connect(DATABASE_PATH)
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()

        # Update or insert Customer List
        cursor.execute("SELECT * FROM 'Customer List' WHERE `Customer ID` = ?", (customer_id,))
        existing = cursor.fetchone()
        if existing:
            cursor.execute("""
                UPDATE 'Customer List'
                SET `Customer Name` = ?, `Address` = ?, `City` = ?, `Province` = ?, `Postal Code` = ?, `Region` = ?
                WHERE `Customer ID` = ?
            """, (customer_name, address, city, province, postal_code, region, customer_id))
        else:
            cursor.execute("""
                INSERT INTO 'Customer List'
                (`Customer ID`, `Customer Name`, `Address`, `City`, `Province`, `Postal Code`, `Region`)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (customer_id, customer_name, address, city, province, postal_code, region))

        # Check if entry already exists in Run Sheet
        cursor.execute("""
            SELECT * FROM 'Run Sheet'
            WHERE `Customer ID` = ? AND `Region` = ?
        """, (customer_id, region))
        run_entry = cursor.fetchone()

        if run_entry:
            new_weight = int(run_entry['Weight']) + weight
            new_skids = int(run_entry['Skids']) + skids
            new_bundles = int(run_entry['Bundles']) + bundles
            new_coils = int(run_entry['Coils']) + coils
            new_closing_time = closing_time if closing_time != run_entry['Closing Time'] else run_entry['Closing Time']

            cursor.execute("""
                UPDATE 'Run Sheet'
                SET Weight = ?, Skids = ?, Bundles = ?, Coils = ?, `Closing Time` = ?, Pickup = ?
                WHERE `Customer ID` = ? AND `Region` = ?
            """, (new_weight, new_skids, new_bundles, new_coils, new_closing_time, pickup, customer_id, region))

        else:
            # Insert new entry
            cursor.execute("""
                INSERT INTO 'Run Sheet'
                (`Customer ID`, `Customer Name`, `Address`, `City`, `Province`, `Postal Code`, `Region`,
                 `Weight`, `Skids`, `Bundles`, `Coils`, `Closing Time`, `Pickup`)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (customer_id, customer_name, address, city, province, postal_code, region,
                  weight, skids, bundles, coils, closing_time, pickup))

        connection.commit()
        connection.close()

        flash(f"✅ Successfully added {customer_name} to {region} region.")
        return redirect(url_for('home.home', customer_id=customer_id))

    # GET mode: Load customer for edit
    if customer_id:
        connection = sqlite3.connect(DATABASE_PATH)
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM 'Customer List' WHERE `Customer ID` = ?", (customer_id,))
        customer_data = cursor.fetchone()
        connection.close()

    return render_template('add_customer.html', customer=customer_data)







# -------------------
# Import Run Sheet
# -------------------

@runsheet_bp.route('/import_run_sheet', methods=['GET'])
def show_import_form():
    return render_template('import_run_sheet.html')
@runsheet_bp.route('/import_run_sheet', methods=['POST'])
def import_run_sheet():
    file = request.files['file']
    if not file:
        flash('No file selected.')
        return redirect(url_for('runsheet.view_run_sheet'))

    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    connection = sqlite3.connect(DATABASE_PATH, timeout=10)
    cursor = connection.cursor()
    cursor.execute("PRAGMA journal_mode=WAL;")

    # Clear both Run Sheet and Drivers before importing
    cursor.execute("DELETE FROM 'Run Sheet'")
    cursor.execute("DELETE FROM Drivers")

    region_blocks = {
        'North Shore': ('A3', 'I18'),
        'Quebec': ('A22', 'I37'),
        'Montreal': ('A41', 'I56'),
        'Ontario': ('A60', 'I75'),
        'Drummond': ('K3', 'S18'),
        'Beauce': ('K22', 'S37'),
        'South Shore': ('K41', 'S56'),
        'Sherbrooke': ('K60', 'S75')
    }

    driver_cells = {
        'North Shore': 'B1',
        'Quebec': 'B20',
        'Montreal': 'B39',
        'Ontario': 'B58',
        'Drummond': 'L1',
        'Beauce': 'L20',
        'South Shore': 'L39',
        'Sherbrooke': 'L58'
    }

    def safe_string(value):
        if value is None:
            return ''
        if isinstance(value, time):
            return value.strftime('%H:%M')
        return str(value).strip()

    def safe_int(value):
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0

    for region, (start_cell, end_cell) in region_blocks.items():
        start_col_letter, start_row = coordinate_from_string(start_cell)
        end_col_letter, end_row = coordinate_from_string(end_cell)
        start_col = column_index_from_string(start_col_letter)
        end_col = column_index_from_string(end_col_letter)

        # Read driver name from specified cell
        driver_cell = driver_cells.get(region)
        driver_name = safe_string(ws[driver_cell].value) if driver_cell else "____________________"
        if not driver_name:
            driver_name = "____________________"

        cursor.execute("INSERT INTO Drivers (Name, Region) VALUES (?, ?)", (driver_name, region))

        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True
        ):
            if not row or all(cell in (None, "", " ") for cell in row):
                continue  # Skip fully blank rows

            row = list(row) + [""] * (9 - len(row))  # Pad short rows

            customer_id = safe_string(row[0])
            customer_name = safe_string(row[1])
            city = safe_string(row[2])
            weight = safe_string(row[3])
            skids = safe_int(row[4])
            bundles = safe_int(row[5])
            coils = safe_int(row[6])
            closing_time = safe_string(row[7])
            pickup = safe_string(row[8])

            cursor.execute('''
                INSERT INTO "Run Sheet" 
                ("Customer ID", "Customer Name", "City", "Weight", "Skids", "Bundles", "Coils", "Closing Time", "Pickup", "Region")
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (customer_id, customer_name, city, weight, skids, bundles, coils, closing_time, pickup, region))

    connection.commit()
    connection.close()

    flash('✅ Successfully imported new Run Sheet!')
    return redirect(url_for('runsheet.view_run_sheet'))

@runsheet_bp.route('/clear_run_sheet')
def clear_run_sheet():
    try:
        connection = sqlite3.connect(DATABASE_PATH, timeout=10)
        cursor = connection.cursor()
        cursor.execute("PRAGMA journal_mode=WAL;")
        cursor.execute("DELETE FROM 'Run Sheet'")
        cursor.execute("DELETE FROM Drivers")
        connection.commit()
        flash('✅ Successfully cleared the Run Sheet.')
    except sqlite3.OperationalError as e:
        flash(f"❌ Failed to clear Run Sheet: {str(e)}")
    finally:
        if connection:
            connection.close()

    return redirect(url_for('runsheet.view_run_sheet'))


# -------------------
# Save and Download Run Sheet
# -------------------
@runsheet_bp.route('/save_run_sheet')
def save_run_sheet():
    try:
        save_run_sheet_to_excel()
        return redirect(url_for('runsheet.download_run_sheet'))
    except Exception as e:
        print("Error saving run sheet:", str(e))
        return "Failed to save run sheet.", 500

@runsheet_bp.route('/download_run_sheet')
def download_run_sheet():
    excel_stream = save_run_sheet_to_excel()

    # Generate next-day filename
    today = datetime.now()
    next_day = today + timedelta(days=3) if today.weekday() == 4 else today + timedelta(days=1)
    weekday_name = next_day.strftime('%A').upper()
    date_string = next_day.strftime('%m.%d.%Y')
    filename = f"{weekday_name} RUN SHEET {date_string}.xlsx"

    return send_file(
        excel_stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# -------------------
# Delete Run Customer
# -------------------
@runsheet_bp.route('/delete_run_customer', methods=['POST'])
def delete_run_customer():
    connection = sqlite3.connect(DATABASE_PATH)
    cursor = connection.cursor()

    id = request.form.get('id')
    customer_id = request.form.get('customer_id')
    region = request.form.get('region')

    if id:
        cursor.execute("DELETE FROM 'Run Sheet' WHERE id = ?", (id,))
    elif customer_id and region:
        cursor.execute("DELETE FROM 'Run Sheet' WHERE `Customer ID` = ? AND `Region` = ?", (customer_id, region))
    else:
        connection.close()
        return "Cannot delete: no valid identifier found.", 400

    connection.commit()
    connection.close()
    return redirect(url_for('runsheet.view_run_sheet'))

# -------------------
# Placeholder for Edit
# -------------------
@runsheet_bp.route('/edit_run_customer/<customer_id>', methods=['GET'])
def edit_run_customer(customer_id):
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM 'Customer List' WHERE `Customer ID` = ?", (customer_id,))
    customer = cursor.fetchone()
    connection.close()

    if not customer:
        flash("Customer not found.")
        return redirect(url_for('runsheet.view_run_sheet'))

    return render_template("add_customer.html", customer=customer, editing=True)


@runsheet_bp.route('/lookup_customer', methods=['POST'])
def lookup_customer():
    data = request.get_json()
    customer_id = data.get('customer_id')

    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    # Search for exact matches AND extended versions (e.g., 1234, 1234A, 1234B)
    cursor.execute("SELECT * FROM 'Customer List' WHERE `Customer ID` LIKE ?", (customer_id + '%',))
    matches = cursor.fetchall()
    connection.close()

    if not matches:
        return jsonify({ "found": False })

    if len(matches) == 1:
        row = matches[0]
        return jsonify({
            "found": True,
            "exact": True,
            "customer": {
                "customer_id": row['Customer ID'],
                "customer_name": row['Customer Name'],
                "address": row['Address'],
                "city": row['City'],
                "province": row['Province'],
                "postal_code": row['Postal Code'],
                "region": row['Region']
            }
        })
    else:
        # Multiple matches, send list back to frontend
        customer_list = [{
            "customer_id": row['Customer ID'],
            "customer_name": row['Customer Name'],
            "address": row['Address'],
            "city": row['City'],
            "province": row['Province'],
            "postal_code": row['Postal Code'],
            "region": row['Region']
        } for row in matches]

        return jsonify({
            "found": True,
            "exact": False,
            "matches": customer_list
        })



@runsheet_bp.route('/print_run_sheet')
def print_run_sheet():
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM 'Run Sheet'")
    runsheet = cursor.fetchall()

    cursor.execute("SELECT * FROM Drivers")
    drivers = cursor.fetchall()
    connection.close()

    regions = ['North Shore', 'Quebec', 'Montreal', 'Ontario', 'Drummond', 'Beauce', 'South Shore', 'Sherbrooke']
    region_data = {region: [] for region in regions}
    region_totals = {region: {"Weight": 0, "Skids": 0, "Bundles": 0, "Coils": 0} for region in regions}
    driver_info = {region: {"name": ""} for region in regions}

    def safe_int(value):
        try:
            return int(float(value))
        except:
            return 0

    for entry in runsheet:
        region = entry['Region']
        if region in region_data:
            region_data[region].append(entry)
            region_totals[region]["Weight"] += safe_int(entry['Weight'])
            region_totals[region]["Skids"] += safe_int(entry['Skids'])
            region_totals[region]["Bundles"] += safe_int(entry['Bundles'])
            region_totals[region]["Coils"] += safe_int(entry['Coils'])

    for row in drivers:
        region = row['Region']
        if region in driver_info:
            driver_info[region]['name'] = row['Name']

    return render_template("print_run_sheet.html", region_data=region_data, region_totals=region_totals, driver_info=driver_info)


@runsheet_bp.route('/update_run_entry', methods=['POST'])
def update_run_entry():
    customer_id = request.form['customer_id']
    region = request.form['region']
    skids = request.form['skids']
    bundles = request.form['bundles']
    coils = request.form['coils']
    weight = request.form['weight']
    closing_time = request.form['closing_time']
    pickup = request.form['pickup']

    connection = sqlite3.connect(DATABASE_PATH)
    cursor = connection.cursor()
    cursor.execute("""
        UPDATE 'Run Sheet'
        SET Skids = ?, Bundles = ?, Coils = ?, Weight = ?, `Closing Time` = ?, Pickup = ?
        WHERE `Customer ID` = ? AND Region = ?
    """, (skids, bundles, coils, weight, closing_time, pickup, customer_id, region))

    connection.commit()
    connection.close()
    flash(f"✅ Updated entry for {customer_id} in {region}")
    return redirect(url_for('runsheet.view_run_sheet'))

@runsheet_bp.route('/statistics', strict_slashes=False)
def view_statistics():
    return redirect('/dash_statistics/')


def get_region_data():
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM 'Run Sheet'")
    runsheet = cursor.fetchall()

    regions = ['North Shore', 'Quebec', 'Montreal', 'Ontario', 'Drummond', 'Beauce', 'South Shore', 'Sherbrooke']
    region_data = {region: [] for region in regions}

    def safe_int(value):
        try:
            return int(float(value))
        except:
            return 0

    for entry in runsheet:
        region = entry['Region']
        if region in region_data:
            region_data[region].append({
                'Weight': entry['Weight'],
                'Skids': entry['Skids'],
                'Bundles': entry['Bundles'],
                'Coils': entry['Coils']
            })

    connection.close()
    return region_data
